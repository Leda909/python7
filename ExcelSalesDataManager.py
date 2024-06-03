import openpyxl
import matplotlib.pyplot as plt

print("Welcome to our sales data manager which uses Excel")

def create_excel_file(filename):
  # Create a new excel file and add in some data
  workbook = openpyxl.Workbook()
  sheet = workbook.active
  sheet.title = "Data"
  
  # Add headers
  headers = ["Months", "Sales"]
  sheet.append(headers)
  
  # Add sample data
  data = [
    ["January", 150],
    ["February", 200],
    ["March", 250],
    ["April", 300],
    ["May", 350],
    ["June", 400],
    ["July", 450],
    ["August", 500],
    ["September", 550],
    ["October", 600],
    ["November", 650],
    ["December", 700]
  ]
  
  for row in data:
      sheet.append(row)
      
  workbook.save(filename)
  print(f"The excel file {filename} was created successfully!")

def read_excel_file(filename):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
        
    return data

def visualise_data(data):
    # Extract months and sales data
    months = [row[0] for row in data[1:]]  # Exclude header
    sales = [row[1] for row in data[1:]]  # Exclude header
    
    plt.figure(figsize=(10, 6))
    plt.plot(months, sales, marker='o', linestyle='-', color='b')
    plt.title('Monthly Sales Data')
    plt.xlabel('Months')
    plt.ylabel('Sales')
    plt.grid(True)
    plt.show()
    
def main():
  filename = 'sales_data.xlsx'
  
  #Create Exccel file
  create_excel_file(filename)
  
  #Read data from excel file
  data = read_excel_file(filename)
  print("This are the data from excel file: ")
  for row in data:
    print(row)
  
  #Visualise the data in a line graph  
  visualise_data(data)
  
if __name__ == "__main__":
  main()
